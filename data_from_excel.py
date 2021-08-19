
#Open Pyxl module
#pip install openpyxl

'''
1- Reading data from excel file
2- Writing data into excel file
3- Data driven test case
'''

import openpyxl
path = 'F:\demo_excel.xlsx'

workbook = openpyxl.load_workbook(path)

#if you have single sheet in your excel u will only run this cmd
sheet = workbook.active

#if you have multiple sheet in excel
#sheet = workbook.get_sheet_by_name('sheet1')

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for i in range(1,rows+1):
    for j in range(1,cols+1):
        print(sheet.cell(row=i,column=j).value,end='           ')
    print()